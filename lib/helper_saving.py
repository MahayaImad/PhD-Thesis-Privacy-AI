
import os
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from datetime import datetime
from openpyxl.drawing.image import Image as ExcelImage
from tensorflow.keras.utils import plot_model
import matplotlib.pyplot as plt
from io import BytesIO
import seaborn as sns

class TrainingReportSaver:
    
    image_path = "model_architecture.png"
    scale = 1
    
    def __init__(self, save_path):
        self.save_path = save_path
        self.wb = Workbook()
        self.metrics_sheet = self.wb.active
        self.metrics_sheet.title = "Metrics"
        self.info_sheet = self.wb.create_sheet("General_Info")
        self.params_sheet = self.wb.create_sheet("Hyperparameters")
        self.arch_sheet = self.wb.create_sheet("Model_Architecture")
    

    def log_metrics(self, train_metrics, val_metrics, f1_score, confusion_matrix,test_accuracy,test_loss):
        num_epochs = len(train_metrics[list(train_metrics.keys())[0]])
        epochs = list(range(1, num_epochs + 1))

        header = ["Epoch"]
        for key in train_metrics:
            header.append(f"Train {key.capitalize()}")
        for key in val_metrics:
            header.append(f"Val {key.capitalize()}")
        self.metrics_sheet.append(header)

        for i in range(num_epochs):
            row = [epochs[i]]
            for key in train_metrics:
                row.append(train_metrics[key][i])
            for key in val_metrics:
                row.append(val_metrics[key][i])
            self.metrics_sheet.append(row)

        i = num_epochs + 1
        c = get_column_letter(len(train_metrics) + len(val_metrics) + 1)
        tab = Table(displayName="Metrics", ref=f"A1:{c}{i}")
        
        style = TableStyleInfo(name="TableStyleLight11", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        self.metrics_sheet.add_table(tab)

        
        for i in range(len(train_metrics) + len(val_metrics) + 1):
            self.metrics_sheet.column_dimensions[get_column_letter(i+1)].width = 13
                    
        metric_names = list(train_metrics.keys())
        header = ["Epoch"] + [f"Train {k}" for k in train_metrics] + [f"Val {k}" for k in val_metrics]

        for i, metric in enumerate(metric_names):
            train_col = header.index(f"Train {metric}") + 1
            val_col = header.index(f"Val {metric}") + 1
            chart_position = f"G{1 + i * 20}"  # Adjust spacing to fit image height

            train_values = [
                self.metrics_sheet.cell(row=row, column=train_col).value
                for row in range(2, 2 + num_epochs)
            ]
            val_values = [
                self.metrics_sheet.cell(row=row, column=val_col).value
                for row in range(2, 2 + num_epochs)
            ]

            self._add_chart_with_plot(
                title=f"{metric.capitalize()} Chart",
                train_values=train_values,
                val_values=val_values,
                position=chart_position
            )    
        # for i, metric in enumerate(metric_names):
        #     train_col = header.index(f"Train {metric}") + 1
        #     val_col = header.index(f"Val {metric}") + 1
        #     chart_position = f"G{1 + i * 15}"
        #     self._add_chart(
        #         title=f"{metric.capitalize()} Chart",
        #         col_indices=[train_col, val_col],
        #         position=chart_position,
        #         num_epochs=num_epochs
        #     )

        i = num_epochs + 1   
        
        self.metrics_sheet.cell(row=i+3, column=1, value="Test Accuracy")
        self.metrics_sheet.cell(row=i+3, column=2, value="Test Loss")
        self.metrics_sheet.cell(row=i+3, column=3, value="F1 Score")
        self.metrics_sheet.cell(row=i+4, column=1, value=test_accuracy)
        self.metrics_sheet.cell(row=i+4, column=2, value=test_loss)
        self.metrics_sheet.cell(row=i+4, column=3, value=f1_score)

        tab2 = Table(displayName="Test_Metrics", ref=f"A{i+3}:C{i+4}")
        tab2.tableStyleInfo = style
        self.metrics_sheet.add_table(tab2)

        cm_start_row = i + 27  
        cm_start_col = 1

        num_classes = len(confusion_matrix)
        self.metrics_sheet.cell(cm_start_row, cm_start_col, value="Confusion Matrix")
        for col in range(num_classes):
            self.metrics_sheet.cell(row=cm_start_row, column=cm_start_col + col + 1, value=f"Pred {col}")

        for row in range(num_classes):
            self.metrics_sheet.cell(row=cm_start_row + row + 1, column=cm_start_col, value=f"True {row}")
            for col in range(num_classes):
                self.metrics_sheet.cell(row=cm_start_row + row + 1, column=cm_start_col + col + 1, value=confusion_matrix[row][col])

        tab3 = Table(displayName="Confusion_Matrix", ref=f"A{cm_start_row}:{get_column_letter(num_classes+1)}{cm_start_row+num_classes}")
        tab3.tableStyleInfo = style
        self.metrics_sheet.add_table(tab3)

    def _add_chart_with_plot(self, title, train_values, val_values, position):
        plt.figure(figsize=(6, 4))
        sns.lineplot(x=range(1, len(train_values) + 1), y=train_values, label='Train')
        sns.lineplot(x=range(1, len(val_values) + 1), y=val_values, label='Validation')
        plt.title(title)
        plt.xlabel('Epoch')
        plt.ylabel(title.split()[0])  
        plt.legend()
        plt.tight_layout()

        # Save plot to a BytesIO buffer
        img_data = BytesIO()
        plt.savefig(img_data, format='png')
        plt.close()
        img_data.seek(0)

        # Insert into Excel
        img = Image(img_data)
        img.anchor = position
        self.metrics_sheet.add_image(img)

    def log_hyperparameters(self, hyperparams):
        self.params_sheet.column_dimensions['A'].width = 25
        self.params_sheet.column_dimensions['B'].width = 30
        self.params_sheet["A1"] = "Hyperparameter"
        self.params_sheet["B1"] = "Value"
        self.params_sheet["B1"].alignment = Alignment(horizontal='center', vertical='center')

        for i, (key, val) in enumerate(hyperparams.items(), start=2):
            self.params_sheet[f"A{i}"] = key
            self.params_sheet[f"B{i}"] = val
            self.params_sheet[f"A{i}"].font = Font(bold=True)
            self.params_sheet[f"B{i}"].alignment = Alignment(horizontal='center', vertical='center')

        i = len(hyperparams) + 1
        tab = Table(displayName="Hyperparameter", ref=f"A1:B{i}")
        
        style = TableStyleInfo(name="TableStyleLight10", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        self.params_sheet.add_table(tab)
        
    def log_experiment_info(self, experiment_info):
        self.info_sheet.column_dimensions['A'].width = 25
        self.info_sheet.column_dimensions['B'].width = 30
        self.info_sheet["A1"] = "Info"
        self.info_sheet["B1"] = "Value"
        self.info_sheet["B1"].alignment = Alignment(horizontal='center', vertical='center')

        for i, (key, val) in enumerate(experiment_info.items(), start=2):
            self.info_sheet[f"A{i}"] = key
            self.info_sheet[f"B{i}"] = val
            self.info_sheet[f"A{i}"].font = Font(bold=True)
            self.info_sheet[f"B{i}"].alignment = Alignment(horizontal='center', vertical='center')

        i = len(experiment_info) + 1
        tab = Table(displayName="experiment_info", ref=f"A1:B{i}")
        
        style = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        self.info_sheet.add_table(tab)
        

    def save(self):
        save_dir = os.path.dirname(self.save_path)
        os.makedirs(save_dir, exist_ok=True)
        self.wb.save(self.save_path)
        print(f"✅ Report saved to: {self.save_path}")

    def _add_chart(self, title, col_indices, position, num_epochs):
        chart = LineChart()
        chart.title = title
        chart.style = 10
        chart.y_axis.title = title.split()[0]
        chart.x_axis.title = "Epoch"
    
        cats = Reference(self.metrics_sheet, min_col=1, min_row=2, max_row=num_epochs + 1)
    
        for col in col_indices:
            data = Reference(self.metrics_sheet, min_col=col, max_col=col, min_row=1, max_row=num_epochs + 1)
            chart.add_data(data, titles_from_data=True)
    
        chart.set_categories(cats)
        col_letter = get_column_letter(col_indices[0])
        chart_position = f"{col_letter}{10 * col_indices[0]}"
        self.metrics_sheet.add_chart(chart, position or chart_position)

    def insert_model_image(self, model):
        try:
            plot_model(model, to_file=self.image_path, show_shapes=True)
            if os.path.exists(self.image_path):
                img = Image(self.image_path)
                img.height = int(img.height * self.scale)
                img.width = int(img.width * self.scale)
                self.arch_sheet.add_image(img, "D3")
            else:
                print(f"⚠️ Image file not found: {self.image_path}")
        except Exception as e:
            print(f"⚠️ Could not generate model plot: {e}")

def save_training(save_path,model,history,hyperparams,experiment_info,f1_score, confusion_matrix,test_accuracy,test_loss):
  history = history.history
  train_metrics = {
      "loss" : history['loss'],
      "acc" : history['accuracy'],
  }

  val_metrics = {
      "loss" : history['val_loss'],
      "acc" : history['val_accuracy'],
  }
  saver = TrainingReportSaver(save_path)
  saver.log_metrics(train_metrics, val_metrics,f1_score, confusion_matrix,test_accuracy,test_loss)
  saver.log_hyperparameters(hyperparams)
  saver.log_experiment_info(experiment_info)
  saver.insert_model_image(model)
  saver.save()